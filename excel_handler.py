import os, re, subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import config

PALETTE = [
    "FCE4EC","F3E5F5","EDE7F6","E8EAF6","E3F2FD",
    "E0F7FA","E0F2F1","E8F5E9","FFFDE7","FFF3E0",
    "FBE9E7","EFEBE9","ECEFF1","F1F8E9","F9FBE7",
    "FFF8E1","FFF3E0","FBE9E7","FCE4EC","F3E5F5",
    "EDE7F6","E8EAF6","E3F2FD","E0F7FA","E0F2F1",
    "E8F5E9","FFFDE7","FFF3E0","FBE9E7","EFEBE9",
]
CATEGORY_FILL_MAP = {}
def get_fill_for_category(cat):
    if cat not in CATEGORY_FILL_MAP:
        idx = len(CATEGORY_FILL_MAP) % len(PALETTE)
        CATEGORY_FILL_MAP[cat] = PatternFill(start_color=PALETTE[idx], end_color=PALETTE[idx], fill_type="solid")
    return CATEGORY_FILL_MAP[cat]

def sanitize_display_name(name):
    return re.sub(r'[^a-zA-Z0-9_]', '', name.replace(" ","_"))

def create_fresh_excel(job_items, course_items, pdf_items, excel_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    default_fieldnames = ["title","link","description","source","category"]
    def add_sheet(ws_name, items):
        ws = wb.create_sheet(title=ws_name)
        for col_idx, field in enumerate(default_fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, item in enumerate(items, 2):
            for col_idx, field in enumerate(default_fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item.get(field, ""))
                if field == "link":
                    cell.hyperlink = item.get("link", "")
                    cell.font = Font(color="0563C1", underline="single")
                if field == "category":
                    cat = item.get("category","")
                    if cat: cell.fill = get_fill_for_category(cat)
                    cell.alignment = Alignment(horizontal="center")
        widths = [60,50,30,25,25]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        last_col = get_column_letter(len(default_fieldnames))
        last_row = len(items)+1
        ws.freeze_panes = "A2"
        table_ref = f"A1:{last_col}{last_row}"
        table_name = sanitize_display_name(ws_name)
        tab = Table(displayName=table_name, ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)
        return ws
    if job_items: add_sheet("Remote Jobs", job_items)
    if course_items: add_sheet("Free Tech Courses", course_items)
    if pdf_items: add_sheet("PDF Resources", pdf_items)
    wb.save(excel_path)
    print(f"New Excel workbook created: {excel_path}")
    subprocess.run(["open", excel_path])

def update_excel_safe(job_items, course_items, pdf_items, excel_path):
    default_fieldnames = ["title","link","description","source","category"]
    if not os.path.exists(excel_path):
        create_fresh_excel(job_items, course_items, pdf_items, excel_path)
        return
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Existing Excel file is corrupted ({e}). Recreating a new one.")
        os.remove(excel_path)
        create_fresh_excel(job_items, course_items, pdf_items, excel_path)
        return

    def append_new_rows_no_table(ws, new_items, fieldnames):
        headers = [cell.value for cell in ws[1]]
        col_map = {name: idx+1 for idx, name in enumerate(headers) if name}
        existing_links = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row)>=2 and row[1]:
                existing_links.add(row[1].strip())
        start_row = ws.max_row + 1
        added = 0
        for item in new_items:
            if item["link"] in existing_links: continue
            added += 1
            row_idx = start_row + added - 1
            for field in fieldnames:
                if field in col_map:
                    col = col_map[field]
                    cell = ws.cell(row=row_idx, column=col, value=item.get(field, ""))
                    if field == "link":
                        cell.hyperlink = item.get("link", "")
                        cell.font = Font(color="0563C1", underline="single")
                    if field == "category":
                        cat = item.get("category","")
                        if cat:
                            cell.fill = get_fill_for_category(cat)
                        cell.alignment = Alignment(horizontal="center")
            existing_links.add(item["link"])
        if added > 0:
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
            print(f"  Added {added} new rows to {ws.title}")
        else:
            print(f"  No new rows for {ws.title}")

    sheet_mapping = {
        "Remote Jobs": (job_items, default_fieldnames),
        "Free Tech Courses": (course_items, default_fieldnames),
        "PDF Resources": (pdf_items, default_fieldnames),
    }
    for sheet_name, (items, fields) in sheet_mapping.items():
        if not items: continue
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            append_new_rows_no_table(ws, items, fields)
        else:
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=1, column=col_idx, value=field).font = Font(bold=True)
            for row_idx, item in enumerate(items, 2):
                for col_idx, field in enumerate(fields, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=item.get(field, ""))
                    if field == "link":
                        cell.hyperlink = item.get("link", "")
                    if field == "category":
                        cat = item.get("category","")
                        if cat: cell.fill = get_fill_for_category(cat)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            print(f"  Created missing sheet: {sheet_name} with {len(items)} items")
    wb.save(excel_path)
    print(f"Updated Excel workbook: {excel_path}")
    subprocess.run(["open", excel_path])
